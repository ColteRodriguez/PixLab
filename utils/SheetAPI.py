import openpyxl

# Methods for spreadsheet manipulation
def locate_sample(spreadsheet_path, name_query, sheet_num):
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook[workbook.sheetnames[sheet_num]]
    # Only search the first row becasue the sample name cant be anywhere else
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value) == name_query:
                return cell.row, cell.column
            
    print(f"Query for {name_query} at location {spreadsheet_path} returned Null. Adding {name_query}...")
    return None, None
        
def get_cell(row, col, spreadsheet_path, sheet_num):
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook[workbook.sheetnames[sheet_num]]
    
    val = sheet.cell(row=row, column=col).value
    workbook.save(spreadsheet_path)
    return val

def change_cell(row, col, spreadsheet_path, val, sheet_num):
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook[workbook.sheetnames[sheet_num]]
    
    sheet.cell(row=row, column=col, value=val)
    workbook.save(spreadsheet_path)
    return val

def update_percent(Pcurr, Tcurr, segment_area, rating, isSubject):
    # Div by 0 error should not occur because total area is always updated 
    # Update the rating
    if rating is not None:
        return ((Pcurr * Tcurr) + (rating * segment_area)) / (Tcurr + segment_area)
    # Update the constituent or feature
    elif isSubject == True:
        return ((Pcurr * Tcurr) + (segment_area)) / (Tcurr + segment_area)
    # Update other constituents whose areas are not to be updated
    else:
        return (Pcurr * Tcurr) / (Tcurr + segment_area)
    
    
def add_new_sample(path, name):
    def find_first_empty_row_header(path):
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[workbook.sheetnames[0]]

        # Start from row 2 since row 1 is the header
        for row in range(2, sheet.max_row + 2):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value is None or cell_value == "":
                return row
        return None  # Return None if no empty cell is found

    row = find_first_empty_row_header(path)
    print(row)
    
    # Add in the name and fill in 0s
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]
    sheet.cell(row=row, column=1, value=name)
    for i in range(2,sheet.max_column + 1):
        sheet.cell(row=row, column=i, value=0)
        
    # Add in the name and fill in 0s
    sheet = workbook[workbook.sheetnames[1]]
    sheet.cell(row=row, column=1, value=name)
    for i in range(2,sheet.max_column + 1):
        sheet.cell(row=row, column=i, value=0)
    workbook.save(path)
    
    
def find_open_cell(filename):
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(filename)
    ws = workbook[workbook.sheetnames[0]]
    
    # Iterate through the cells in the first row in reverse order
    for cell in reversed(ws[1]):
        if cell.value is not None:
            return cell.row, cell.column + 1  # Return the cell's coordinate (e.g., 'Z1')
    
    return None  # Return None if all cells are empty

from openpyxl.utils import get_column_letter

def set_zeros(filename, col):
    workbook = openpyxl.load_workbook(filename)
    ws = workbook[workbook.sheetnames[0]]
    
    i = 1
    for cell in ws[get_column_letter(1)]:
        if i!=1: # Don't set the column header
            if cell.value or cell.value is not None: # If there is a sample
                ws.cell(row=i, column=col, value=0)
        i+=1
    
    workbook.save(filename)
    
# Dear god here we go, lets keep this organized
def update_spreadsheet(polygon_area, constituent, image, altertion_score, path):
    sheet_path = path
    
    # 1.0.1: Query for the sample name (This will always exist if segmented images are used)
    image_name = image[image.index('Img'):image.index('Img')+7]
    sample_row, sample_col = locate_sample(sheet_path, image_name, 0)
    
    
    # 1.1: Query for if the constituent exists. If no, add it
    const_row, const_col = locate_sample(sheet_path, constituent, 0)
    if const_row == None or const_col == None:
        const_row, const_col = find_open_cell(sheet_path)
        change_cell(const_row, const_col, sheet_path, constituent, 0) # Add new constituent
        
        # Set the new constituent val to 0 for all samples
        set_zeros(sheet_path, const_col)
        
    # 1.2: get the previous total mapped area value
    prev_total_area = get_cell(sample_row, 2, sheet_path, 0)
    
    # 1.2.1: Update the total mapped area
    change_cell(sample_row, 2, sheet_path, int(prev_total_area) + int(polygon_area), sheet_num = 0)
    
    ##### Continue
    # 2.0: Update the Constituent percentage distrib
    for i in range(4, const_col+1):
        rating = None
        isSubject = False
        
        # Update params
        if get_cell(1, i, sheet_path, 0) == constituent:
            isSubject = True
        old_value = get_cell(sample_row, i, sheet_path, sheet_num = 0)
        new_value = update_percent(old_value, prev_total_area, polygon_area, rating, isSubject)
        change_cell(sample_row, i, sheet_path, new_value, sheet_num = 0)
    
    # 5. Update the Alteration Score
    r, c = sample_row, 4
    old_value = get_cell(r, c, sheet_path, sheet_num = 0)
    new_value = update_percent(old_value, prev_total_area, polygon_area, rating = altertion_score, isSubject = False)
    change_cell(r, c, sheet_path, new_value, sheet_num = 0)
    
    
    
